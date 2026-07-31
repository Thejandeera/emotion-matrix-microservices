import sys
import subprocess
import time
import os

# Auto-install pandas and openpyxl if missing in environment
try:
    import pandas as pd
    import openpyxl
except ModuleNotFoundError:
    print("[Setup] Missing 'pandas' or 'openpyxl'. Installing dependencies...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd
    import openpyxl

import torch
from transformers import pipeline
from openpyxl.styles import PatternFill, Font

DATASET_PATH = "dataset.xlsx"

def main():
    print("=" * 85)
    print("     DATASET EMOTION EVALUATION & COMPARISON: RoBERTa vs ModernBERT")
    print("=" * 85)

    if not os.path.exists(DATASET_PATH):
        print(f"❌ Error: Dataset file '{DATASET_PATH}' not found in workspace root.")
        return

    print(f"\n[1/4] Reading dataset from '{DATASET_PATH}'...")
    try:
        df = pd.read_excel(DATASET_PATH)
        print(f"      ✓ Successfully loaded {len(df)} rows. Columns: {list(df.columns)}")
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return

    if "text" not in df.columns:
        print("❌ Error: Required column 'text' not found in dataset.xlsx!")
        return

    has_actual_emotion = "actual_emotion" in df.columns

    print("\n[2/4] Loading HuggingFace Pipelines...")
    print("      -> Loading RoBERTa ('SamLowe/roberta-base-go_emotions')...")
    try:
        roberta_pipe = pipeline("text-classification", model="SamLowe/roberta-base-go_emotions")
        print("      ✓ RoBERTa loaded.")
    except Exception as e:
        print(f"❌ Failed to load RoBERTa: {e}")
        return

    print("      -> Loading ModernBERT ('cirimus/modernbert-base-go-emotions')...")
    try:
        modernbert_pipe = pipeline("text-classification", model="cirimus/modernbert-base-go-emotions")
        print("      ✓ ModernBERT loaded.")
    except Exception as e:
        print(f"❌ Failed to load ModernBERT: {e}")
        return

    # Warmup models once
    dummy_text = "Hello world this is a test text."
    with torch.inference_mode():
        _ = roberta_pipe(dummy_text, truncation=True, max_length=256)
        _ = modernbert_pipe(dummy_text, truncation=True, max_length=256)

    print(f"\n[3/4] Evaluating {len(df)} text samples across both models...")

    roberta_emotions = []
    roberta_confidences = []
    roberta_times = []
    
    modernbert_emotions = []
    modernbert_confidences = []
    modernbert_times = []
    
    models_match = []

    mismatch_count = 0
    roberta_correct_count = 0
    modernbert_correct_count = 0

    for idx, row in df.iterrows():
        text_val = str(row["text"]).strip() if pd.notna(row["text"]) else ""
        actual_val = str(row["actual_emotion"]).strip().lower() if has_actual_emotion and pd.notna(row["actual_emotion"]) else None

        if not text_val:
            roberta_emotions.append("N/A")
            roberta_confidences.append(0.0)
            roberta_times.append(0.0)
            modernbert_emotions.append("N/A")
            modernbert_confidences.append(0.0)
            modernbert_times.append(0.0)
            models_match.append("MATCH")
            continue

        # RoBERTa evaluation with timing
        r_start = time.perf_counter()
        with torch.inference_mode():
            r_res = roberta_pipe(text_val, truncation=True, max_length=256)[0]
        r_time_ms = round((time.perf_counter() - r_start) * 1000, 2)
        r_label = r_res["label"]
        r_score = round(float(r_res["score"]), 4)

        # ModernBERT evaluation with timing
        m_start = time.perf_counter()
        with torch.inference_mode():
            m_res = modernbert_pipe(text_val, truncation=True, max_length=256)[0]
        m_time_ms = round((time.perf_counter() - m_start) * 1000, 2)
        m_label = m_res["label"]
        m_score = round(float(m_res["score"]), 4)

        is_match = (r_label.lower() == m_label.lower())
        if not is_match:
            mismatch_count += 1

        if actual_val:
            if r_label.lower() == actual_val:
                roberta_correct_count += 1
            if m_label.lower() == actual_val:
                modernbert_correct_count += 1

        roberta_emotions.append(r_label)
        roberta_confidences.append(r_score)
        roberta_times.append(r_time_ms)
        
        modernbert_emotions.append(m_label)
        modernbert_confidences.append(m_score)
        modernbert_times.append(m_time_ms)
        
        models_match.append("MATCH" if is_match else "MISMATCH")

        status_symbol = "✓ MATCH" if is_match else "❌ MISMATCH"
        actual_str = f" | Actual: {actual_val}" if actual_val else ""
        print(f"Row {idx+1:02d}/{len(df)}: '{text_val[:30]}...'{actual_str} | RoBERTa: {r_label} ({r_score*100:.1f}%, {r_time_ms}ms) | ModernBERT: {m_label} ({m_score*100:.1f}%, {m_time_ms}ms) | {status_symbol}")

    # Build updated DataFrame with exact desired columns
    columns_order = ["text"]
    if has_actual_emotion:
        columns_order.append("actual_emotion")

    df["roberta_emotion"] = roberta_emotions
    df["roberta_confidence"] = roberta_confidences
    df["roberta_execution_time_ms"] = roberta_times
    
    df["modernbert_emotion"] = modernbert_emotions
    df["modernbert_confidence"] = modernbert_confidences
    df["modernbert_execution_time_ms"] = modernbert_times
    
    df["models_match"] = models_match

    # Save DataFrame to dataset.xlsx
    df.to_excel(DATASET_PATH, index=False)
    print(f"\n[4/4] Applying Excel styling for mismatched predictions in '{DATASET_PATH}'...")

    # Highlight mismatched rows using openpyxl
    wb = openpyxl.load_workbook(DATASET_PATH)
    ws = wb.active

    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    mismatch_font = Font(color="9C0006", bold=True)

    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    
    rob_col = header_map.get("roberta_emotion")
    mod_col = header_map.get("modernbert_emotion")
    match_col = header_map.get("models_match")

    for row_idx in range(2, ws.max_row + 1):
        rob_val = ws.cell(row=row_idx, column=rob_col).value
        mod_val = ws.cell(row=row_idx, column=mod_col).value
        
        if rob_val and mod_val and str(rob_val).lower() != str(mod_val).lower():
            if rob_col:
                ws.cell(row=row_idx, column=rob_col).fill = mismatch_fill
                ws.cell(row=row_idx, column=rob_col).font = mismatch_font
            if mod_col:
                ws.cell(row=row_idx, column=mod_col).fill = mismatch_fill
                ws.cell(row=row_idx, column=mod_col).font = mismatch_font
            if match_col:
                ws.cell(row=row_idx, column=match_col).fill = mismatch_fill
                ws.cell(row=row_idx, column=match_col).font = mismatch_font

    wb.save(DATASET_PATH)

    avg_rob_time = sum(roberta_times) / len(roberta_times) if roberta_times else 0
    avg_mod_time = sum(modernbert_times) / len(modernbert_times) if modernbert_times else 0

    print("\n" + "=" * 85)
    print("                     EVALUATION COMPLETED SUCCESSFULLY!")
    print("=" * 85)
    print(f"📊 Total Rows Evaluated            : {len(df)}")
    print(f"🎯 Model Prediction Agreement Rate  : {((len(df) - mismatch_count) / len(df)) * 100:.2f}% ({len(df) - mismatch_count}/{len(df)})")
    print(f"⚠️ Mismatched Model Predictions   : {mismatch_count} (highlighted in soft red in Excel)")
    
    if has_actual_emotion:
        print("-" * 85)
        print(f"🏆 RoBERTa Accuracy vs Actual      : {(roberta_correct_count / len(df)) * 100:.2f}% ({roberta_correct_count}/{len(df)})")
        print(f"🏆 ModernBERT Accuracy vs Actual   : {(modernbert_correct_count / len(df)) * 100:.2f}% ({modernbert_correct_count}/{len(df)})")

    print("-" * 85)
    print(f"⚡ Average RoBERTa Execution Time  : {avg_rob_time:.2f} ms")
    print(f"⚡ Average ModernBERT Execution Time: {avg_mod_time:.2f} ms")
    if avg_rob_time > 0 and avg_mod_time > 0:
        faster_model = "ModernBERT" if avg_mod_time < avg_rob_time else "RoBERTa"
        diff_ms = abs(avg_rob_time - avg_mod_time)
        print(f"🚀 Speed Advantage                  : {faster_model} is {diff_ms:.2f} ms faster per sample on average")

    print("=" * 85)
    print(f"💾 Saved updated results to Excel   : {os.path.abspath(DATASET_PATH)}")
    print("=" * 85)

if __name__ == "__main__":
    main()
